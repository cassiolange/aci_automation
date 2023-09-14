import argparse
import time
import os
import sys
import logging
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import TableStyleInfo, Table, TableFormula
from openpyxl.styles import Color, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension

import yaml

def data_validation_indirect(worksheet, source_column_name, yaml_source_file, cell_position):
    formula_format = 'INDIRECT(${column}2)'
    position = get_column_position(yaml_source_file=yaml_source_file, worksheet_name=worksheet.title, column_name=source_column_name)
    formula = formula_format.format(column=position['column_letter'])
    worksheet = create_data_validation(formula=formula, worksheet=worksheet, cell_position=cell_position, check_quotes=False)
    return worksheet

def set_cell_color(worksheet, color, cell_position):
    match color['type']:
        case 'rgb':
            cell_color = Color(type='rgb', rgb= color['rgb'])
        case 'theme':
            cell_color = Color(type='theme', theme=color['theme'], tint=color['tint'])
    if cell_color:
        worksheet[cell_position].fill = PatternFill(patternType='solid', fgColor=cell_color)
    return worksheet

def set_tab_color(worksheet, color):
    match color['type']:
        case 'rgb':
            tab_color = Color(type='rgb', rgb= color['rgb'])
            worksheet.sheet_properties.tabColor = tab_color
        case 'theme':
            tab_color = Color(type='theme', theme=color['theme'], tint=color['tint'])
            worksheet.sheet_properties.tabColor = tab_color

    return worksheet
def data_validation_enabled_disabled(worksheet, cell_position, max_cell=None):
    worksheet = create_data_validation(formula='enabled,disabled', worksheet=worksheet, cell_position=cell_position, max_cell=max_cell)
    return worksheet
def data_validation_yes_no(worksheet, cell_position, default_value = None, max_cell=None):
    worksheet = create_data_validation(formula='yes,no', worksheet=worksheet, cell_position=cell_position, max_cell=max_cell)
    if default_value:
        worksheet[cell_position] = default_value
    return worksheet

def data_validation_status(worksheet, cell_position, max_cell=None):
    worksheet = create_data_validation(formula='present,absent,ignored,deployed', worksheet=worksheet, cell_position=cell_position, max_cell=max_cell)
    return worksheet

def create_data_validation(formula, worksheet, cell_position, dv_type='list', check_quotes=True, max_cell=None):
    if check_quotes:
        if formula.startswith('\"') == False:
            formula = '\"'+formula
        if formula.endswith('\"') == False:
            formula = formula+'\"'
    dv = DataValidation(type=dv_type, formula1=formula, allowBlank=True)
    worksheet.add_data_validation(dv)
    if max_cell:
        dv.add(cell_position+':'+max_cell)
    else:
        dv.add(cell_position)
    return worksheet

def create_defined_names(workbook, name, formula):
    defined_name = DefinedName(name, attr_text=formula)
    workbook.defined_names.add(defined_name)
    return workbook
def data_validation_from_worksheet(worksheet, source_worksheet_name, source_column_name, yaml_source_file, cell_position, max_cell=None):
    defined_name = source_worksheet_name+'_'+source_column_name
    defined_name_formula = f'{source_worksheet_name}[{source_column_name}]'
    if defined_name not in worksheet.defined_names:
        worksheet = create_defined_names(workbook=worksheet, name=defined_name, formula=defined_name_formula)
    worksheet = create_data_validation(formula=defined_name, worksheet=worksheet, cell_position=cell_position, check_quotes=False, max_cell=max_cell)
    return worksheet

def create_formula(formula, column_number, worksheet, cell_position):
    current_table_name = str(worksheet.title)
    worksheet[cell_position] = formula
    column_formula = TableFormula()
    column_formula.attr_text = formula
    worksheet.tables[current_table_name].tableColumns[column_number-1].calculatedColumnFormula = column_formula
    #worksheet.tables[current_table_name].tableColumns[column_number - 1].dataDxfId = 155
    worksheet[cell_position].fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF00B0F0'), bgColor=Color(indexed=64))
    return worksheet

def process_sheet_fields(ws, input_yaml, sheet_name):
    for column,value in enumerate(input_yaml[sheet_name], start=1):
        ws.cell(row=1, column=column, value=str(value))
        current_column_letter = get_column_letter(column)
        current_cell_position = current_column_letter + '2'
        max_row = ws.max_row
        if max_row == 1:
            max_cell = None
        else:
            max_cell = current_column_letter + str(ws.max_row)
        if isinstance(input_yaml[sheet_name][value], dict):
            try:
                match input_yaml[sheet_name][value]['type']:
                    case 'data_validation_static_list':
                        if 'check_quotes' in input_yaml[sheet_name][value]:
                            check_quotes = input_yaml[sheet_name][value]['check_quotes']
                        else:
                            check_quotes = True
                        ws = create_data_validation(worksheet=ws, cell_position=current_cell_position, formula=input_yaml[sheet_name][value]['formula'], max_cell=max_cell, check_quotes=check_quotes)
                    case 'data_validation_from_worksheet':
                        ws = data_validation_from_worksheet(worksheet=ws, source_worksheet_name=input_yaml[sheet_name][value]['source'], source_column_name=input_yaml[sheet_name][value]['column'], yaml_source_file=input_yaml, cell_position=current_cell_position, max_cell=max_cell)
                    case 'xlookup':
                        xlookup_format = '=IF(ISBLANK({current_table}[[#This Row],[{lookup_field}]]),"",_xlfn.XLOOKUP({current_table}[[#This Row],[{lookup_field}]],{source_table}[{match_field}],{source_table}[{return_field}]))'
                        if 'match_field' in input_yaml[sheet_name][value] and 'return_field' in input_yaml[sheet_name][value]:
                            xlookup = xlookup_format.format(current_table=sheet_name, match_field=str(input_yaml[sheet_name][value]['match_field']), lookup_field=str(input_yaml[sheet_name][value]['lookup_field']), source_table=input_yaml[sheet_name][value]['source_table'], return_field=input_yaml[sheet_name][value]['return_field'])
                        else:
                            xlookup = xlookup_format.format(current_table=sheet_name, match_field=str(input_yaml[sheet_name][value]['lookup_field']), lookup_field=str(input_yaml[sheet_name][value]['lookup_field']), source_table=input_yaml[sheet_name][value]['source_table'], return_field=str(value))
                        max_row = ws.max_row
                        if max_row == 1:
                            max_row = 2
                        for row in range(2, max_row + 1):
                            ws = create_formula(formula=xlookup, column_number=column, worksheet=ws, cell_position=current_column_letter+str(row))
                    case 'data_validation_indirect':
                        ws = data_validation_indirect(worksheet=ws, source_column_name=input_yaml[sheet_name][value]['source'], yaml_source_file=input_yaml, cell_position=current_cell_position)
                    case 'formula_static':
                        max_row = ws.max_row
                        if max_row == 1:
                            max_row = 2
                        for row in range(2, max_row + 1):
                            ws = create_formula(formula=input_yaml[sheet_name][value]['formula'], column_number=column, worksheet=ws,cell_position=current_column_letter+str(row))


            except:
                sys.exit('Worksheet %s, column %s' % (sheet_name,value))
        elif isinstance(input_yaml[sheet_name][value], str):
            match input_yaml[sheet_name][value]:
                case 'yes_no':
                    ws = data_validation_yes_no(worksheet=ws, cell_position=current_cell_position, max_cell=max_cell)
                case 'status':
                    ws = data_validation_status(worksheet=ws, cell_position=current_cell_position, max_cell=max_cell)
                case 'enabled_disabled':
                    ws = data_validation_enabled_disabled(worksheet=ws, cell_position=current_cell_position, max_cell=max_cell)
def create_table(worksheet, table_name, columns, start_position='A1', end_row='2'):
    table_style = create_table_style()
    letter_last_column = get_column_letter(len(columns))
    # letter_last_column = get_column_letter(worksheet.max_column)
    size = str(start_position)+':' + letter_last_column + str(end_row)
    table = Table(displayName=table_name, ref=size)
    table.tableStyleInfo = table_style
    table._initialise_columns()
    if start_position=='A1':
        for column, column_name, column_number in zip(table.tableColumns, columns, range(1,len(columns)+1)):
            column.name = str(column_name)
            current_column_letter = get_column_letter(column_number)
            worksheet.cell(row=1, column=column_number, value=str(column_name))
            # worksheet.column_dimensions[current_column_letter].width = len(column_name)+4
            worksheet.column_dimensions[current_column_letter].bestFit = True

    worksheet.add_table(table)
    return worksheet
def sanitize_category(sheet):

    if 'ignore_build_task' in sheet:
        sheet.pop('ignore_build_task')
        ignore_build_taks = True
    else:
        ignore_build_taks = False
    if 'tab_color' in sheet:
        color = sheet['tab_color']
        sheet.pop('tab_color')
    else:
        color= False
    if 'category' in sheet:
        category = sheet['category']
        sheet.pop('category')
    else:
        category = None
    if 'playbook' in sheet:
        playbook = sheet['playbook']
        sheet.pop('playbook')
    else:
        playbook = None

    return sheet, ignore_build_taks, color, category, playbook

def sanitize_columns(sheet):
    if 'build_tasks_description' in sheet:
        build_tasks_description = sheet['build_tasks_description']
        sheet.pop('build_tasks_description')
    else:
        build_tasks_description = None

    return sheet, build_tasks_description

def open_yaml_file(input_yaml_file):
    try:
        logging.info(f'Opening input file {input_yaml_file}')
        input_text_file = open(input_yaml_file)
    except:
        logging.info(f'Input not exist {input_yaml_file}. Exiting')
        sys.exit(-1)

    input_yaml = yaml.safe_load(input_text_file)
    input_text_file.close()
    return input_yaml
def fix_tables_yaml_based(workbook, input_yaml, fix_tab_color=False, update_all_fields=False):
    new_build_tasks = {}
    yes_no_build_task_color = {
        'type': 'theme',
        'theme': 0,
        'tint': -0.1499984740745262
    }
    logging.info(f'Checking the sheets')
    number_of_sheets = 0
    sheets_ignore_build_task = 0
    for category in input_yaml:
        sheets, ignore_build_tasks, color, category_name, playbook = sanitize_category(input_yaml[category])
        if ignore_build_tasks:
            sheets_ignore_build_task = sheets_ignore_build_task + len(input_yaml[category])
        for sheet_name in sheets:
            columns, build_tasks_description = sanitize_columns(sheets[sheet_name])
            table_modified = False
            tables_to_delete = {}
            '''Checking if the sheet exist, if not create '''
            if sheet_name not in workbook:
                logging.info(f'Sheet {sheet_name} not present, adding')
                worksheet = workbook.create_sheet(title=sheet_name, index=number_of_sheets+list(sheets).index(sheet_name))
                if color:
                    worksheet = set_tab_color(worksheet=worksheet, color=color)
                if not ignore_build_tasks:
                    if workbook['build_tasks']['B4'].value == 'Node Registration':
                        workbook['build_tasks']['B4'].value = 'Node Provisioning'
                    if workbook['build_tasks']['B5'].value == 'Node Addressing':
                        workbook['build_tasks'].delete_rows(5)
                    if workbook['build_tasks']['B6'].value == 'Add Node to Firmware group':
                        workbook['build_tasks'].delete_rows(6)
                    if workbook['build_tasks']['B6'].value == 'Add Node to Maintenance group':
                        workbook['build_tasks'].delete_rows(6)
                    build_tasks_new_row_position = number_of_sheets + list(sheets).index(sheet_name) - sheets_ignore_build_task + 1
                    if workbook['build_tasks']['C' + str(build_tasks_new_row_position)].value:
                        workbook['build_tasks'].insert_rows(build_tasks_new_row_position)
                    data_validation_yes_no(worksheet=workbook['build_tasks'], cell_position='A' + str(build_tasks_new_row_position), default_value='no')
                    if build_tasks_description:
                        workbook['build_tasks']['B' + str(build_tasks_new_row_position)].value = build_tasks_description
                    else:
                        workbook['build_tasks'].cell(column=2, row=build_tasks_new_row_position,value=sheet_name.capitalize().replace('_', ' '))
                    link = '#' + sheet_name + '!A1'
                    workbook['build_tasks']['C' + str(build_tasks_new_row_position)].hyperlink = link
                    workbook['build_tasks']['C' + str(build_tasks_new_row_position)].value = sheet_name
                    workbook['build_tasks']['C' + str(build_tasks_new_row_position)].style = "Hyperlink"
                    workbook['build_tasks'].tables['build_tasks'].ref = 'A1:E' + str(workbook['build_tasks'].max_row)
                    cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),  bottom=Side(style='thin'))
                    workbook['build_tasks']['A' + str(build_tasks_new_row_position)].border = cell_border
                    set_cell_color(worksheet=workbook['build_tasks'], color=yes_no_build_task_color, cell_position='A' + str(build_tasks_new_row_position))
                    if category_name:
                        workbook['build_tasks']['D' + str(build_tasks_new_row_position)].value = category_name
                    if color:
                        set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='B' + str(build_tasks_new_row_position))
                        set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='C' + str(build_tasks_new_row_position))
                        set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='D' + str(build_tasks_new_row_position))
                worksheet = create_table(worksheet=worksheet, table_name=sheet_name, columns=columns)
                table_modified = True
                # worksheet = process_sheet_fields(worksheet=worksheet, sheet_name=sheet_name, input_yaml=input_yaml)


            '''Checking the tables name'''
            logging.info(f'Checking Tables Name')
            if workbook[sheet_name].tables.values():
                for table in workbook[sheet_name].tables.values():
                    if table.name != sheet_name:
                        tables_to_delete.update({sheet_name : table.name})
            else:
                create_table(worksheet=workbook[sheet_name], table_name=sheet_name, columns=columns,end_row=workbook[sheet_name].max_row)
            if tables_to_delete:
                del workbook[sheet_name].tables[table.name]
                create_table(worksheet=workbook[sheet_name], table_name=sheet_name, columns=columns, end_row=workbook[sheet_name].max_row)
                table_modified = True


            logging.info(f'Checking the columns in sheets {sheet_name}')

            for column in columns:
                if column not in workbook[sheet_name].tables[sheet_name].column_names:
                    logging.info(f'Column {column} not present in {sheet_name}, adding')
                    new_column_postion = list(columns).index(column) + 1
                    workbook[sheet_name].insert_cols(new_column_postion)
                    workbook[sheet_name].cell(row=1, column=new_column_postion, value=column.upper())
                    table_modified = True

            '''Process the sheet field'''
            if table_modified or update_all_fields:
                '''removing old data validation'''
                for count in range(workbook[sheet_name].data_validations.count):
                    del workbook[sheet_name].data_validations.dataValidation[0]
                del workbook[sheet_name].tables[sheet_name]
                max_row = workbook[sheet_name].max_row
                if max_row == 1:
                    max_row = 2
                create_table(worksheet=workbook[sheet_name], table_name=sheet_name, columns=columns, start_position='A1', end_row=max_row)
                process_sheet_fields(ws=workbook[sheet_name], sheet_name=sheet_name, input_yaml=sheets)

            if fix_tab_color:
                worksheet = set_tab_color(worksheet=workbook[sheet_name], color=color)
        number_of_sheets = number_of_sheets + len(input_yaml[category])
    return workbook


def create_table_style():
    table_style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    return table_style

def create_backup(file, backup_dir='../spreadsheets/backup'):
    if not os.path.isdir(backup_dir):
        logging.info(f'Creating backup folder {backup_dir}')
        os.mkdir(backup_dir)
    if backup_dir.endswith('/') == False:
        backup_dir = backup_dir+'/'
    destination_file = backup_dir+file.split('/')[len(file.split('/'))-1].split('.')[0]+'_'+str(time.strftime("%d_%m_%y_%H_%M_%S_%s"))+'.'+file.split('/')[len(file.split('/'))-1].split('.')[1]
    logging.info(f'Creating file backup at {destination_file}')
    os.system(f'cp {file} {destination_file}')
def open_excel_file(excel_file):
    if not os.path.isfile(excel_file):
        logging.info(f'Excel file {excel_file} not exist. Exiting.')
        exit(-1)

    logging.info('Opening File %s', excel_file)
    workbook = load_workbook(excel_file)
    return workbook


def build_columns_dict(workbook, worksheet, table=None):
    columns_name = {}
    if not table:
        table=worksheet

    for column in workbook[worksheet].tables[table].tableColumns:
        columns_name.update({column.name: len(columns_name)})
    return columns_name


def create_data_source_sheet(input_yaml, workbook, sheet_name='data_source'):
    for value in input_yaml[sheet_name]:
        match value:
            case 'defined_names':
                for defined_name in input_yaml[sheet_name]['defined_names']:
                    workbook = create_defined_names(workbook=workbook, name=defined_name['name'] , formula=defined_name['formula'])

    return workbook

def resize_columns(workbook, sheet_name=None):
    if sheet_name:
        for cells in workbook[sheet_name].iter_cols():
            column_width = 0
            for cell in cells:
                if cell.value:
                    column_width = max(column_width, len(cell.value))
            workbook[sheet_name].column_dimensions[cell.column_letter].width = column_width
    else:
        print('to be implemented')
    return workbook
def create_build_tasks(workbook, input_yaml):
    yes_no_build_task_color = {
        'type': 'theme',
        'theme': 0,
        'tint': -0.1499984740745262
    }
    worksheet = workbook.create_sheet(title='build_tasks', index=0)
    build_tasks_columns, build_tasks_description = sanitize_columns(input_yaml['build_tasks']['build_tasks'])
    worksheet = set_tab_color(worksheet=worksheet, color=input_yaml['build_tasks']['tab_color'])
    ColumnDimension(worksheet, bestFit=True)
    column_width = {}
    for index_column,column in enumerate(build_tasks_columns, 1):
        workbook['build_tasks'].cell(row=1, column=index_column, value=column)

    for category in input_yaml:
        sheets, ignore_build_tasks, color, category_name, playbook = sanitize_category(input_yaml[category])
        if not ignore_build_tasks:
            for sheet_name in sheets:
                columns, build_tasks_description = sanitize_columns(sheets[sheet_name])
                current_row = workbook['build_tasks'].max_row + 1
                data_validation_yes_no(worksheet=workbook['build_tasks'], cell_position='A' + str(current_row), default_value='no')
                if build_tasks_description:
                    workbook['build_tasks']['B' + str(current_row)].value = build_tasks_description
                else:
                    workbook['build_tasks'].cell(column=2, row=current_row, value=sheet_name.capitalize().replace('_', ' '))
                link = '#' + sheet_name + '!A1'
                workbook['build_tasks']['C' + str(current_row)].hyperlink = link
                workbook['build_tasks']['C' + str(current_row)].value = sheet_name
                workbook['build_tasks']['C' + str(current_row)].style = "Hyperlink"
                cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                workbook['build_tasks']['A' + str(current_row)].border = cell_border
                set_cell_color(worksheet=workbook['build_tasks'], color=yes_no_build_task_color, cell_position='A' + str(current_row))
                if category_name:
                    workbook['build_tasks']['D' + str(current_row)].value = category_name
                if color:
                    set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='B' + str(current_row))
                    set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='C' + str(current_row))
                    set_cell_color(worksheet=workbook['build_tasks'], color=color, cell_position='D' + str(current_row))

    worksheet = create_table(worksheet=worksheet, table_name='build_tasks', columns=build_tasks_columns, end_row=workbook['build_tasks'].max_row)
    workbook = resize_columns(workbook=workbook, sheet_name='build_tasks')


    return workbook

def reorder_tabs(sheet):
    input_file_yaml = '../input_yaml/aci.yaml'
    input_yaml = open_yaml_file(input_yaml_file=input_file_yaml)
    sheet_tab_position = 0
    for category in input_yaml:
        sheets, ignore_build_tasks, color, category_name, playbook = sanitize_category(input_yaml[category])
        if sheet.title in sheets:
            return sheet_tab_position + +list(sheets).index(sheet.title)
        sheet_tab_position += len(sheets)
    return sheet_tab_position

def type_and_version(workbook):
    logging.info('Trying to determine the automation type and Excel version')
    if 'data_source' in workbook and 'version' in workbook['data_source'].tables:
        table_size = str(workbook['data_source'].tables['version'].ref).split(':')
        automation_type = workbook['data_source'][table_size[0].replace('1', '2')].value
        excel_version = workbook['data_source'][table_size[1]].value
        logging.info(f'Automation Type is {automation_type}')
        logging.info(f'Excel Version {excel_version}')
        return automation_type, excel_version
    else:
        logging.info('Unable to determine the Automation Type, exiting.')
        exit(0)

def main():
    parser = argparse.ArgumentParser(description="Script to Update and Fix the Excel File.")
    parser.add_argument("excel_file_location", type=str, help="Location of the ACI Excel File EX:../spreedsheets/aci_build.xlsx")
    parser.add_argument("--rearrange_tabs", default=False, action="store_true", help="Rearrange the tabs in the Excel Sheet.")
    parser.add_argument("--recreate_build_tasks", default=False, action="store_true", help="Recreate the Build Tasks sheet")
    parser.add_argument("--create_backup", default=True, action="store_true", help="Make a copy of the Excel File")
    parser.add_argument("--input_yaml_folder", type=str, default='../input_yaml/', help=argparse.SUPPRESS)
    parser.add_argument("--fix_tab_color", default=False, action="store_true", help="Correct the color of the Excel tabs")
    parser.add_argument("--update_all", default=False, action="store_true", help="Update all sheets in the Excel File")
    parser.add_argument("--output_file_location", type=str, default='', help="Output file, if left blank, equals the input file.")


    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format='%(asctime)s:%(levelname)s:%(message)s', handlers=[logging.FileHandler('excel_file.log'), logging.StreamHandler()])
    if args.create_backup:
        create_backup(file=args.excel_file_location)
    workbook = open_excel_file(args.excel_file_location)
    automation_type, excel_version = type_and_version(workbook=workbook)


    input_yaml = open_yaml_file(input_yaml_file=args.input_yaml_folder+automation_type+'.yaml')
    if 'data_source' in input_yaml:
        workbook = create_data_source_sheet(input_yaml=input_yaml, workbook=workbook)
        input_yaml.pop('data_source')
    workbook = fix_tables_yaml_based(workbook=workbook, input_yaml=input_yaml, fix_tab_color=args.fix_tab_color, update_all_fields=args.update_all)

    if args.recreate_build_tasks:
        del workbook['build_tasks']
        input_yaml = open_yaml_file(input_yaml_file=args.input_yaml_file)
        if 'data_source' in input_yaml:
            input_yaml.pop('data_source')
        workbook = create_build_tasks(workbook=workbook, input_yaml=input_yaml)

    if args.rearrange_tabs:
        workbook._sheets.sort(key=reorder_tabs)

    if args.output_file_location:
        workbook.save(args.output_file_location)
    else:
        workbook.save(args.excel_file_location)



if __name__ == '__main__':
    main()